'''
    File name: newsScrape.py
    Python Version: 3.*

    Description: Generate tweets from excel file
    Usage: generate('AquíNecesitamos.xlsx', 'output.json')
'''

import json
import re
from openpyxl import load_workbook

src_sheet = ["URGENCIAS Y SOLICITUDES POR ZON"]
sheet_index = 0;
MAX_TWEET_CHARACTERS = 138
DATA_MIN_ROW = 6
DATA_MAX_COL = 9
INFO_HASH_TAG = "#infoverificada19S"

# Data label
URGENT_LEVEL = {"alto":"URGE", "alta":"URGE", "medio":"SeNecesita", "media":"SeNecesita", "bajo":"SeNecesita", "baja":"SeNecesita"}

#-------------------------------------------------------
# Main
#-------------------------------------------------------
def generate(src_filename, dst_filename):
    tweetList = [];
    wb = load_workbook(src_filename)
    ws = wb[src_sheet[sheet_index]]
    for row in ws.iter_rows(min_row=DATA_MIN_ROW, max_col=DATA_MAX_COL):
        row_data = []
        for cell in row:
            row_data.append(cell.value);
        tweet = generateTweet(row_data)
        tweetList.append(convertTweetToJson(tweet))
    jsonData = createJson(tweetList)
    print("Generated: " + str(len(tweetList)) + " tweets")
    saveFile(jsonData, dst_filename)
    print("---Finish saving generated text to: " + dst_filename + "---")

# Generate tweet from row data
# [0] Urgent Level [1] NEED BRIGADISTS [3] MOST IMPORTANT REQUIRED
# [4] ADMITTED [5] NOT REQUIRED [6] ADDRESS
# [7] ZONE [8] DETAIL/SOURCE [9] UPGRADE
def generateTweet(row):
    tweet = ""

    # Time
    if (row[8] is not None):
        time = (row[8])[5:] # remove year
        tweet += time

    # Urgent level
    if (row[0] is not None):
        urgent = row[0].lower()
        urgent = urgent.rstrip()
        urgent = urgent.lstrip()
        tweet += " " + URGENT_LEVEL[urgent]+" "

##    # Need Brigadists
##    needHelp = (row[1].lower() == "si") or (row[1].lower() == "sí")
##    if (needHelp):
##        tweet += ", NECESITAN BRIGADISTAS"

    # Hash Tag
    tweet += INFO_HASH_TAG

    # Address
    if (row[5] is not None):
        address = row[5]
        address = getAddress(address)
        tweet += " Rescate en " + address

    # Zone
    if (row[6] is not None):
        zone = row[6]
        tweet += " " + zone

    return tweet

# Parse address and return address text
def getAddress(address):
    address = str(address)
    formatText = '=HYPERLINK'
    hasLink = (address[0:len(formatText)]) == formatText
    if (hasLink):
        addressTmp = re.sub( "HYPERLINK\(\".+\"," , "" , address)
        address = addressTmp[2:len(addressTmp)-2]
        return address
    else:
        return address

def checkTweetLength(text):
    return (len(text) < MAX_TWEET_CHARACTERS)

# Convert string to JSON format
def convertTweetToJson(tweet):
    data = {"text": tweet}
    return json.dumps(data)

# Create JSON from the list
def createJson(tweetList):
    output = "["
    for i in range(len(tweetList)):
        if (i == len(tweetList)-1):
            output += tweetList[i]
        else:
            output += tweetList[i] + ", "
    output += "]"
    return output

# Save data to file
def saveFile(data, filename):
    file = open(filename, 'w')
    file.write(data)
    file.close
